import pandas as pd
from openpyxl import load_workbook
wb = load_workbook("C:\Vsuet\Train.xlsx")

sheet = wb.get_sheet_by_name('Лист1')

matr = [[] for i in range(6)]

row = 5
column = 5

for i in range(1, 6):
    for j in range(1, 22):
        print(i, sheet.cell(row=j, column=i).value)
        matr[i].append(sheet.cell(row=j, column=i).value)



for i in matr:
    print("\n")
    for j in i:
        print(j, end="\t")
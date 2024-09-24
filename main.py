import numpy as np
import pandas as pd
from openpyxl import load_workbook


wb = load_workbook("C:\Vsuet\Train.xlsx")

sheet = wb.get_sheet_by_name('Лист1')

matr = [[] for i in range(6)]

matr2 = [[] for i in range(7)]


for i in range(2, 6):
    for j in range(2, 48):
        print(i, sheet.cell(row=j, column=i).value)
        matr[i].append(sheet.cell(row=j, column=i).value)
count = 0
for i in range(10, 15):
    for j in range(2, 23):
        print(count, sheet.cell(row=j, column=i).value)
        matr2[count].append(sheet.cell(row=j, column=i).value)
    count += 1

for i in range(len(matr2)):

    for j in range(len(matr2[i])):
        print(matr2[i][j], end="\t")
    print(end="\n")

data_in = [[] for i in range(len(matr[2]))]

data_in2 = [[] for i in range(len(matr2[2]))]

for i in range(len(matr)):
    for j in range(len(matr[i])):
        data_in[j].append(matr[i][j])

for i in range(len(matr2)):
    for j in range(len(matr2[i])):
        data_in2[j].append(matr2[i][j])

for i in data_in2:
    print("\n")
    for j in i:
        print(j, end="\t")

def sigmoid(x):

    return 1 / (1 + np.exp(-x))


def deriv_sigmoid(x):
    fx = sigmoid(x)
    return fx * (1 - fx)


def mse_loss(y_true, y_pred):
    return ((y_true - y_pred) ** 2).mean()


class OurNeuralNetwork:

    def __init__(self):
        # Вес
        self.w1 = np.random.normal()
        self.w2 = np.random.normal()
        self.w3 = np.random.normal()
        self.w4 = np.random.normal()
        self.w5 = np.random.normal()
        self.w6 = np.random.normal()
        self.w7 = np.random.normal()
        self.w8 = np.random.normal()
        self.w9 = np.random.normal()
        self.w10 = np.random.normal()
        self.w11 = np.random.normal()
        self.w12 = np.random.normal()

        # Смещения
        self.b1 = np.random.normal()
        self.b2 = np.random.normal()
        self.b3 = np.random.normal()
        self.b4 = np.random.normal()
    def print(self):
        print(self.w1, self.w2, self.w3, self.w4, self.w5, self.w6, self.w7, self.w8,
              self.w9, self.w10, self.w11, self.w12, self.b1, self.b1, self.b2, self.b3, self.b4)


    def feedforward(self, x):
        # x является массивом numpy с двумя элементами
        h1 = sigmoid(self.w1 * x[0] + self.w2 * x[1] + self.w3 * x[2] + self.b1)
        h2 = sigmoid(self.w4 * x[0] + self.w5 * x[1] + self.w6 * x[2] + self.b2)
        h3 = sigmoid(self.w7 * x[0] + self.w8 * x[1] + self.w9 * x[2] + self.b3)

        o1 = sigmoid(self.w10 * h1 + self.w11 * h2 + self.w12 * h3 + self.b4)
        return o1

    def end_result(self):
        y_preds = np.apply_along_axis(self.feedforward, 1, data_isk)

        for i in range(len(matr2[0])):
            if y_preds[i] < (sum(y_preds)/len(y_preds)):
                print(matr2[0][i], " - Незачёт")
            else:
                print(matr2[0][i], " - Зачёт")
        return y_preds

    def train(self, data, all_y_trues):

        learn_rate = 0.1
        epochs = 10000

        for epoch in range(epochs):
            for x, y_true in zip(data, all_y_trues):

                sum_h1 = self.w1 * x[0] + self.w2 * x[1] + self.w3 * x[2] + self.b1
                h1 = sigmoid(sum_h1)

                sum_h2 = self.w4 * x[0] + self.w5 * x[1] + self.w6 * x[2] + self.b2
                h2 = sigmoid(sum_h2)

                sum_h3 = self.w7 * x[0] + self.w8 * x[1] + self.w9 * x[2] + self.b3
                h3 = sigmoid(sum_h3)

                sum_o1 = self.w10 * h1 + self.w11 * h2 + self.w12 * h3 + self.b4
                o1 = sigmoid(sum_o1)
                y_pred = o1


                d_L_d_ypred = -2 * (y_true - y_pred)

                # Нейрон o1
                d_ypred_d_w10 = h1 * deriv_sigmoid(sum_o1)
                d_ypred_d_w11 = h2 * deriv_sigmoid(sum_o1)
                d_ypred_d_w12 = h3 * deriv_sigmoid(sum_o1)
                d_ypred_d_b4 = deriv_sigmoid(sum_o1)

                d_ypred_d_h1 = self.w10 * deriv_sigmoid(sum_o1)
                d_ypred_d_h2 = self.w11 * deriv_sigmoid(sum_o1)
                d_ypred_d_h3 = self.w12 * deriv_sigmoid(sum_o1)


                d_h1_d_w1 = x[0] * deriv_sigmoid(sum_h1)
                d_h1_d_w2 = x[1] * deriv_sigmoid(sum_h1)
                d_h1_d_w3 = x[2] * deriv_sigmoid(sum_h1)
                d_h1_d_b1 = deriv_sigmoid(sum_h1)


                d_h2_d_w4 = x[0] * deriv_sigmoid(sum_h2)
                d_h2_d_w5 = x[1] * deriv_sigmoid(sum_h2)
                d_h2_d_w6 = x[2] * deriv_sigmoid(sum_h2)
                d_h2_d_b2 = deriv_sigmoid(sum_h2)


                d_h3_d_w7 = x[0] * deriv_sigmoid(sum_h3)
                d_h3_d_w8 = x[1] * deriv_sigmoid(sum_h3)
                d_h3_d_w9 = x[2] * deriv_sigmoid(sum_h3)
                d_h3_d_b3 = deriv_sigmoid(sum_h3)


                self.w1 -= learn_rate * d_L_d_ypred * d_ypred_d_h1 * d_h1_d_w1
                self.w2 -= learn_rate * d_L_d_ypred * d_ypred_d_h1 * d_h1_d_w2
                self.w3 -= learn_rate * d_L_d_ypred * d_ypred_d_h1 * d_h1_d_w3
                self.b1 -= learn_rate * d_L_d_ypred * d_ypred_d_h1 * d_h1_d_b1

                self.w4 -= learn_rate * d_L_d_ypred * d_ypred_d_h2 * d_h2_d_w4
                self.w5 -= learn_rate * d_L_d_ypred * d_ypred_d_h2 * d_h2_d_w5
                self.w6 -= learn_rate * d_L_d_ypred * d_ypred_d_h2 * d_h2_d_w6
                self.b2 -= learn_rate * d_L_d_ypred * d_ypred_d_h2 * d_h2_d_b2


                self.w7 -= learn_rate * d_L_d_ypred * d_ypred_d_h3 * d_h3_d_w7
                self.w8 -= learn_rate * d_L_d_ypred * d_ypred_d_h3 * d_h3_d_w8
                self.w9 -= learn_rate * d_L_d_ypred * d_ypred_d_h3 * d_h3_d_w9
                self.b3 -= learn_rate * d_L_d_ypred * d_ypred_d_h3 * d_h3_d_b3


                self.w6 -= learn_rate * d_L_d_ypred * d_ypred_d_w10
                self.w7 -= learn_rate * d_L_d_ypred * d_ypred_d_w11
                self.w8 -= learn_rate * d_L_d_ypred * d_ypred_d_w12
                self.b4 -= learn_rate * d_L_d_ypred * d_ypred_d_b4


            if epoch % 10 == 0:
                y_preds = np.apply_along_axis(self.feedforward, 1, data)
                loss = mse_loss(all_y_trues, y_preds)
                print("Epoch %d loss: %.3f" % (epoch, loss))



data = [data_in[i][:-1] for i in range(len(data_in))]
print(len(data))
data_isk = [data_in2[i][1:-1:] for i in range(len(data_in2))]
all_y_trues = matr[-1]
print(all_y_trues)
print(data_isk)

network = OurNeuralNetwork()
network.train(data, all_y_trues)
network.print()
print(network.end_result())